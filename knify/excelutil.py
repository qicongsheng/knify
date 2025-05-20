#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author: qicongsheng
import json
import os
import re
from typing import Callable

import xlrd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from knify import listutil
from knify import logger
from knify import objutil


class Header:
    def __init__(self, index: int, name: str | None,
                 transformer: Callable[[object], object] = None):
        self.index = index
        self.name = name
        self.transformer = transformer


class HeaderBuilder:
    def __init__(self):
        self.default_transformer = None
        self.headers = []

    def set_default_transformer(self,
                                transformer: Callable[[object], object] = None) -> object:
        self.default_transformer = transformer
        return self

    def set_names(self, names: list[str] = None) -> object:
        for index_, name in enumerate(names):
            self.headers.append(Header(index_, name, self.default_transformer))
        return self

    def set_transformer(self, name: str,
                        transformer: Callable[[object], object] = None) -> object:
        for header in self.headers:
            if name == header.name:
                header.transformer = transformer
        return self

    def append(self, index: int, name: str | None,
               transformer: Callable[[object], object] = None) -> object:
        target_index = objutil.default_if_none(index, len(self.headers))
        target_transformer = objutil.default_if_none(transformer,
                                                     self.default_transformer)
        self.headers.append(Header(target_index, name, target_transformer))
        return self

    def to_headers(self) -> list[Header]:
        return self.headers


def read_excel(file_path: str, sheet: str | int | None = 0,
               headers: list[Header] | None = None, start_row: int = 1,
               header_row: int = 0) -> list[object]:
    results = []
    workbook = load_workbook(filename=file_path)
    sheet_ = workbook[sheet] if isinstance(sheet, str) else workbook[
        workbook.sheetnames[sheet]]
    headers_ = [cell.value for cell in sheet_[header_row + 1]]
    for row_idx, row in enumerate(sheet_.rows):
        if row_idx < start_row:
            continue
        result = {}
        for header_idx, header_ in enumerate(headers_):
            # 没有传入headers,使用默认header
            if listutil.is_empty(headers):
                result[header_] = row[header_idx].value
            # 传入了headers
            else:
                header = listutil.find_first(
                    list(filter(lambda h_: h_.index == header_idx, headers)))
                if header is None:
                    continue
                else:
                    col_name = objutil.default_if_none(header.name, header_)
                    cell_value = row[header_idx].value
                    cell_value = cell_value if header.transformer is None else header.transformer(
                        row[header_idx].value)
                    result[col_name] = cell_value
        if objutil.has_keys(result):
            results.append(result)
    return results


def read_headers(file_path: str, sheet: str | int | None = 0,
                 header_row: int = 0):
    workbook = load_workbook(filename=file_path)
    sheet_ = workbook[sheet] if isinstance(sheet, str) else workbook[
        workbook.sheetnames[sheet]]
    return [cell.value for cell in sheet_[header_row + 1]]


def load_excel_data(file_path, sheet_index):
    """通用Excel加载函数，支持xls和xlsx格式"""
    if file_path.endswith('.xls'):
        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_index(sheet_index)
        headers = sheet.row_values(0)
        data = [sheet.row_values(i) for i in range(1, sheet.nrows)]
        return headers, data, None, sheet.name  # xls格式不获取列宽
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        sheet = wb[wb.sheetnames[sheet_index]]
        headers = [cell.value for cell in next(sheet.iter_rows())]
        data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]

        # 获取列宽信息
        column_widths = []
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            column_dim = sheet.column_dimensions.get(col_letter)
            if column_dim and column_dim.width is not None:
                column_widths.append(column_dim.width)
            else:
                column_widths.append(8.43)  # Excel默认列宽
        return headers, data, column_widths, wb.sheetnames[sheet_index]


def compare(file1_path, file2_path, output_path, key_column, sheet_index=[0],
            file1_alias="文件1", file2_alias="文件2"):
    for sheet_idx in sheet_index:
        compare_(file1_path, file2_path, output_path, key_column, sheet_idx, file1_alias, file2_alias)


def compare_(file1_path, file2_path, output_path, key_column, sheet_index=0,
             file1_alias="文件1", file2_alias="文件2"):
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00',
                              fill_type='solid')
    header_fill = PatternFill(start_color='AFEEEE', end_color='AFEEEE',
                              fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    logger.info(f"开始处理sheet:{sheet_index}")
    # 加载数据和列宽
    headers1, data1_rows, col_widths1, sheet_name = load_excel_data(file1_path, sheet_index)
    headers2, data2_rows, _, sheet_name = load_excel_data(file2_path, sheet_index)

    if headers1 != headers2:
        header_diff = []
        for h_idx, h_value in enumerate(headers1):
            if h_value != headers2[h_idx]:
                header_diff.append(h_value)
        raise ValueError("两个文件的表头不一致:%s" % header_diff)
    if key_column not in headers1:
        raise ValueError(f"主键列 {key_column} 不存在")

    key_index = headers1.index(key_column)
    other_headers = [h for h in headers1 if h != key_column]

    # 构建数据字典
    def build_data_dict(data_rows):
        return {row[key_index]: row for row in data_rows}

    data1 = build_data_dict(data1_rows)
    data2 = build_data_dict(data2_rows)

    # 创建结果工作簿
    if os.path.exists(output_path):
        result_wb = load_workbook(output_path)
    else:
        result_wb = Workbook()
        result_wb.remove(result_wb.active)
    result_ws = result_wb.create_sheet(title=sheet_name, index=sheet_index)

    # ========== 构建表头 ==========
    # 第一行结构：主键列 + 合并列名
    header_row1 = [key_column] + [h for h in other_headers for _ in (0, 0)]
    result_ws.append(header_row1)

    # 第二行结构：空主键 + 文件标识
    header_row2 = [""] + [f for h in other_headers for f in
                          (file1_alias, file2_alias)]
    result_ws.append(header_row2)

    # 合并主键列单元格（纵向合并）
    result_ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    result_ws.cell(1, 1).alignment = Alignment(horizontal='center',
                                               vertical='center')

    # 合并其他列单元格（横向合并）
    col_pos = 2
    for h in other_headers:
        result_ws.merge_cells(
            start_row=1,
            start_column=col_pos,
            end_row=1,
            end_column=col_pos + 1
        )
        result_ws.cell(1, col_pos).value = h
        col_pos += 2

    # 设置表头样式
    for cell in result_ws[1]:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    for cell in result_ws[2]:
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # ====== 冻结窗格和筛选 ======
    result_ws.freeze_panes = 'B3'  # 冻结第一列和前两行
    result_ws.auto_filter.ref = f"A2:{get_column_letter(result_ws.max_column)}2"  # 第二行添加筛选

    # ========== 数据对比处理 ==========
    all_keys = sorted(set(data1.keys()) | set(data2.keys()))

    for key_idx, key in enumerate(all_keys):
        if key_idx % 100 == 0:
            logger.info(f"对比进度:{key_idx}/{len(all_keys)}")
        if key_idx == len(all_keys) - 1:
            logger.info(f"对比进度:{len(all_keys)}/{len(all_keys)}")
        row1 = data1.get(key)
        row2 = data2.get(key)

        # 构建结果行
        result_row = [key]
        for h in other_headers:
            idx = headers1.index(h)
            val1 = row1[idx] if row1 else None
            val2 = row2[idx] if row2 else None
            result_row.extend([val1, val2])

        result_ws.append(result_row)
        current_row = result_ws.max_row

        # 设置数据行边框
        for col in range(1, result_ws.max_column + 1):
            result_ws.cell(row=current_row, column=col).border = thin_border

        # 判断行存在情况
        exists1 = key in data1
        exists2 = key in data2

        # 标记差异逻辑
        if exists1 and exists2:
            # 逐列比较差异
            for col_idx, h in enumerate(other_headers, 1):
                idx = headers1.index(h)
                val1 = None if row1[idx] == '' else row1[idx]
                val2 = None if row2[idx] == '' else row2[idx]
                if val1 != val2:
                    target_col = 1 + col_idx * 2
                    result_ws.cell(current_row, target_col).fill = yellow_fill
        else:
            # 整行标黄逻辑
            for col_idx, h in enumerate(other_headers, 1):
                target_col = 1 + col_idx * 2
                if exists1 and not exists2:  # 只存在文件1
                    result_ws.cell(current_row, target_col).fill = yellow_fill
                elif exists2 and not exists1:  # 只存在文件2
                    result_ws.cell(current_row, target_col).fill = yellow_fill

    # ========== 设置列宽 ==========
    if col_widths1:
        # 设置主键列宽
        key_width = col_widths1[key_index]
        result_ws.column_dimensions['A'].width = key_width

        # 设置其他列宽
        for idx, width in enumerate(col_widths1):
            if headers1[idx] == key_column:
                continue
            # 计算结果文件列位置
            pos_in_other = other_headers.index(headers1[idx])
            result_col = 2 + pos_in_other * 2
            # 设置两列宽度
            result_ws.column_dimensions[get_column_letter(result_col)].width = width
            result_ws.column_dimensions[
                get_column_letter(result_col + 1)].width = width

    result_wb.save(output_path)


def json_file_to_excel(json_file, excel_file, skip_keys=None, sort_headers=True):
    """
    将JSON文件中的数据转换为Excel文件
    :param json_file: JSON文件的路径（例如：'data.json'）
    :param excel_file: 输出的Excel文件路径（例如：'output.xlsx'）
    """
    # 从JSON文件中读取数据
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)  # 解析JSON数据
        json_to_excel(data, excel_file, skip_keys, sort_headers)


def json_to_excel(json_data, excel_file, skip_keys=None, sort_headers=True):
    if skip_keys is None:
        skip_keys = set()  # 如果没有提供 skip_keys，默认为空集合
    else:
        skip_keys = set(skip_keys)  # 将列表转换为集合，便于快速查找

    # 创建一个新的工作簿和工作表
    workbook = Workbook()
    sheet = workbook.active

    # 收集所有可能的键（合并所有对象的键）
    all_keys = set()
    for item in json_data:
        all_keys.update(item.keys())  # 将每个对象的键添加到集合中

    # 过滤掉需要跳过的键
    headers = [key for key in all_keys if key not in skip_keys]

    # 对表头进行排序（如果 sort_headers 为 True）
    if sort_headers:
        headers = sorted(headers)  # 默认按字母顺序升序排序

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    def clean_string(value):
        """
        清理字符串中的非法字符
        """
        if isinstance(value, str):
            # 过滤掉 ASCII 控制字符（0x00-0x1F）和非法 Unicode 字符
            return re.sub(r'[\x00-\x1F\x7F]', '', value)
        return value

    # 写入数据
    for row_num, item in enumerate(json_data, 2):
        for col_num, header in enumerate(headers, 1):
            # 如果当前对象没有该键，则写入空值
            value = item.get(header, "")
            cleaned_value = clean_string(value)  # 清理非法字符
            sheet.cell(row=row_num, column=col_num, value=cleaned_value)

    # 保存Excel文件
    workbook.save(excel_file)
    print(f"Excel文件已保存为 {excel_file}")


def process_data(excel_file, process_func, filter_func=None, preprocess_func=None, postprocess_func=None, header_row=1,
                 sheet_index=0):
    """
    通用的数据处理工具，支持从Excel文件中读取数据并进行处理。
    :param excel_file: Excel文件路径
    :param process_func: 数据处理函数，用于生成最终结果
    :param filter_func: 过滤函数，用于判断哪些数据行需要跳过（可选）
    :param preprocess_func: 预处理函数，用于对列数据进行处理（可选）
    :param postprocess_func: 后置处理函数，用于对生成的结果进行处理（可选）
    :param header_row: Excel文件中表头的行号，默认为1（openpyxl的行号从1开始）
    :param sheet_index: 工作表的索引，默认为0（第一个工作表）
    :return: 处理后的结果列表
    """
    # 判断文件格式
    file_ext = os.path.splitext(excel_file)[1].lower()

    if file_ext == '.xlsx':
        # 使用openpyxl读取.xlsx文件
        workbook = load_workbook(excel_file)
        sheet = workbook.worksheets[sheet_index]
        rows = sheet.iter_rows(min_row=header_row + 1, values_only=True)
        header = [cell.value for cell in sheet[header_row]]
    elif file_ext == '.xls':
        # 使用xlrd读取.xls文件
        workbook = xlrd.open_workbook(excel_file)
        sheet = workbook.sheet_by_index(sheet_index)
        rows = [sheet.row_values(row) for row in range(header_row, sheet.nrows)]
        header = sheet.row_values(header_row - 1)
    else:
        raise ValueError("Unsupported file format. Only .xls and .xlsx are supported.")

    # 初始化结果列表
    results = []

    # 遍历每一行数据
    for row in rows:
        # 将行数据与表头组合为字典
        row_data = {header[i]: row[i] for i in range(len(header))}

        # 如果提供了预处理函数，则对列数据进行处理
        if preprocess_func:
            row_data = preprocess_func(row_data)

        # 如果提供了过滤函数，并且过滤函数返回True，则跳过该行
        if filter_func and filter_func(row_data):
            continue

        # 调用数据处理函数生成结果
        result = process_func(row_data)

        # 如果提供了后置处理函数，则对生成的结果进行处理
        if postprocess_func:
            result = postprocess_func(result, row_data)

        # 将生成的结果添加到列表中
        results.append(result)

    return results
